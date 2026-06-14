#!/usr/bin/env python3
"""
Script principal pour le traitement des SMS
- Concaténation des fichiers Brutes avec dédoublonnage
- Création de fichiers par MSISDN du TP
- Routing des SMS dans les onglets correspondants
"""

import os
import sys
import shutil
from datetime import datetime

# Vérification des dépendances
try:
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"ERREUR: Dépendance manquante - {e}")
    print("Installation requise: pip install pandas openpyxl")
    sys.exit(1)

# Styles pour la coloration
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Configuration
BRUTES_DIR = "Brutes"
CONCAT_DIR = "Concat"
ARCHIVE_DIR = "Brutes_archive"
SORTIE_DIR = "Sortie"
ERROR_DIR = "Error"
REFERENCE_DIR = "Reference"
REFERENCE_FILE = os.path.join(REFERENCE_DIR, "msisdn_reference_v1.1.csv")
TP_FILE = "TP/TP_test.xlsx"
ERROR_LOG_FILE = None  # Sera défini lors de l'initialisation


def ensure_directories():
    """Créer les répertoires nécessaires"""
    os.makedirs(BRUTES_DIR, exist_ok=True)
    os.makedirs(CONCAT_DIR, exist_ok=True)
    os.makedirs(ARCHIVE_DIR, exist_ok=True)
    os.makedirs(SORTIE_DIR, exist_ok=True)
    os.makedirs(REFERENCE_DIR, exist_ok=True)
    os.makedirs(ERROR_DIR, exist_ok=True)


def init_error_logging():
    """Initialiser le fichier de log des erreurs"""
    global ERROR_LOG_FILE
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    ERROR_LOG_FILE = os.path.join(ERROR_DIR, f"errorlog_{ts}.txt")
    
    # Écrire l'en-tête
    with open(ERROR_LOG_FILE, 'w', encoding='utf-8') as f:
        f.write("=" * 60 + "\n")
        f.write("FICHIER DE LOG DES ERREURS\n")
        f.write("=" * 60 + "\n")
        f.write(f"Date de création: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 60 + "\n\n")
    
    return ERROR_LOG_FILE


def log_error(message, error_type="ERREUR", traceback_info=""):
    """Logger une erreur dans le fichier de log"""
    global ERROR_LOG_FILE
    
    if ERROR_LOG_FILE is None:
        init_error_logging()
    
    try:
        with open(ERROR_LOG_FILE, 'a', encoding='utf-8') as f:
            f.write(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [{error_type}] {message}\n")
            if traceback_info:
                f.write(f"  Détails: {traceback_info}\n")
            f.write("\n")
    except Exception:
        # Si on ne peut pas écrire dans le log, afficher à la console
        print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] [{error_type}] {message}")
        if traceback_info:
            print(f"  Détails: {traceback_info}")


def save_error_sms(error_df, timestamp):
    """Sauvegarder les SMS en erreur dans un fichier Excel"""
    try:
        if error_df is None or error_df.empty:
            return None
        
        error_file = os.path.join(ERROR_DIR, f"SMS_error_{timestamp}.xlsx")
        error_df.to_excel(error_file, index=False)
        print(f"SUCCESS: Fichier d'erreur SMS créé - {error_file}")
        log_error(f"Fichier SMS_error créé: {error_file} avec {len(error_df)} lignes en erreur")
        return error_file
    except Exception as e:
        log_error(f"Échec de la sauvegarde des SMS en erreur: {e}")
        return None


def load_file(filepath):
    """Charger un fichier CSV ou Excel"""
    try:
        if filepath.endswith('.csv'):
            return pd.read_csv(filepath)
        elif filepath.endswith('.xlsx') or filepath.endswith('.xls'):
            return pd.read_excel(filepath)
        else:
            print(f"AVERTISSEMENT: Format non supporté - {filepath}")
            return None
    except Exception as e:
        print(f"ERREUR: Échec du chargement de {filepath} - {e}")
        return None


def load_all_brutes():
    """Charger tous les fichiers du dossier Brutes"""
    frames = []
    
    if not os.path.exists(BRUTES_DIR):
        print(f"ERREUR: Dossier {BRUTES_DIR} introuvable")
        return pd.DataFrame()
    
    files = os.listdir(BRUTES_DIR)
    if not files:
        print(f"AVERTISSEMENT: Dossier {BRUTES_DIR} vide")
        return pd.DataFrame()
    
    for filename in files:
        filepath = os.path.join(BRUTES_DIR, filename)
        if os.path.isfile(filepath):
            df = load_file(filepath)
            if df is not None:
                frames.append(df)
    
    if not frames:
        print("ERREUR: Aucun fichier valide trouvé dans Brutes")
        return pd.DataFrame()
    
    return pd.concat(frames, ignore_index=True)


def build_concat():
    """Créer le fichier concaténé avec dédoublonnage"""
    try:
        df = load_all_brutes()
        
        if df.empty:
            print("ERREUR: DataFrame vide après chargement")
            return None
        
        # Formater les données pour éviter le format scientifique
        df = format_dataframe(df)
        
        # Dédoublonnage
        initial_count = len(df)
        df = df.drop_duplicates()
        final_count = len(df)
        
        if initial_count != final_count:
            print(f"Dédoublonnage: {initial_count - final_count} doublons supprimés")
        
        # Génération du nom de fichier avec timestamp
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_file = os.path.join(CONCAT_DIR, f"SMS_Concat_{ts}.xlsx")
        
        # Sauvegarde
        df.to_excel(out_file, index=False)
        print(f"SUCCESS: Fichier concaténé créé - {out_file}")
        
        return out_file
        
    except Exception as e:
        print(f"ERREUR: Échec de la création du fichier concaténé - {e}")
        return None


def archive_brutes():
    """Déplacer les fichiers Brutes vers l'archive"""
    try:
        if not os.path.exists(BRUTES_DIR):
            print(f"AVERTISSEMENT: Dossier {BRUTES_DIR} introuvable, rien à archiver")
            return True
        
        files = os.listdir(BRUTES_DIR)
        if not files:
            print(f"AVERTISSEMENT: Dossier {BRUTES_DIR} vide, rien à archiver")
            return True
        
        for filename in files:
            src = os.path.join(BRUTES_DIR, filename)
            dst = os.path.join(ARCHIVE_DIR, filename)
            
            # Gérer les conflits de noms
            counter = 1
            while os.path.exists(dst):
                name, ext = os.path.splitext(filename)
                dst = os.path.join(ARCHIVE_DIR, f"{name}_{counter}{ext}")
                counter += 1
            
            shutil.move(src, dst)
            print(f"Archivé: {filename} -> {dst}")
        
        print("SUCCESS: Tous les fichiers Brutes archivés")
        return True
        
    except Exception as e:
        print(f"ERREUR: Échec de l'archivage - {e}")
        return False


def load_tp_file(tp_path=TP_FILE):
    """Charger le fichier TP"""
    try:
        if not os.path.exists(tp_path):
            print(f"ERREUR: Fichier TP introuvable - {tp_path}")
            return None
        
        tp_df = pd.read_excel(tp_path, header=0)
        
        # Vérifier les colonnes requises
        # Le fichier peut avoir des colonnes numérotées (0, 1, 2, ...) ou nommées
        if 0 in tp_df.columns:
            # Renommer les colonnes basées sur shared strings
            column_mapping = {
                0: 'Identité',
                1: 'Nationalité', 
                2: 'MSISDN',
                3: 'IMSI',
                4: 'IMEI',
                5: 'Niveau',
                6: 'Lien'
            }
            tp_df = tp_df.rename(columns=column_mapping)
        
        required_cols = ['Identité', 'MSISDN']
        for col in required_cols:
            if col not in tp_df.columns:
                print(f"ERREUR: Colonne manquante dans TP - {col}")
                print(f"Colonnes disponibles: {tp_df.columns.tolist()}")
                return None
        
        # Nettoyer les données: convertir MSISDN en string et supprimer les lignes vides
        tp_df['MSISDN'] = tp_df['MSISDN'].astype(str)
        tp_df = tp_df.dropna(subset=['MSISDN'])
        
        print(f"SUCCESS: Fichier TP chargé ({len(tp_df)} lignes)")
        return tp_df
        
    except Exception as e:
        print(f"ERREUR: Échec du chargement du fichier TP - {e}")
        return None


def create_tp_sheets(tp_df, output_file):
    """Créer un fichier Excel avec un onglet par MSISDN du TP + onglet Accueil"""
    try:
        wb = Workbook()
        
        # Supprimer la feuille par défaut
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        sheet_map = {}
        
        for _, row in tp_df.iterrows():
            sheet_name = f"{row['Identité']} - {row['MSISDN']}"
            
            # Excel limite à 31 caractères
            sheet_name = sheet_name[:31]
            
            # Gérer les doublons de noms d'onglets
            original_name = sheet_name
            counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name[:25]}_{counter}"
                counter += 1
            
            ws = wb.create_sheet(title=sheet_name)
            
            # Écrire l'en-tête
            headers = ["ActionDate", "MSISDN APT", "MSISDN APE", "Preview", "Traduction"]
            ws.append(headers)
            
            sheet_map[row["MSISDN"]] = sheet_name
        
        wb.save(output_file)
        print(f"SUCCESS: Fichier avec onglets TP créé - {output_file}")
        return sheet_map
        
    except Exception as e:
        print(f"ERREUR: Échec de la création des onglets TP - {e}")
        return None


def create_pb_format_sheet(wb, concat_df, ref_df):
    """Créer l'onglet PB_Format avec les MSISDN bruts et formatés
    
    Args:
        wb: Workbook openpyxl
        concat_df: DataFrame concaténé des SMS
        ref_df: DataFrame du fichier de référence
    """
    try:
        # Extraire tous les MSISDN uniques des colonnes APT et APE
        all_msisdns = set()
        
        if 'MSISDN APT' in concat_df.columns:
            all_msisdns.update(concat_df['MSISDN APT'].astype(str).unique())
        if 'MSISDN APE' in concat_df.columns:
            all_msisdns.update(concat_df['MSISDN APE'].astype(str).unique())
        
        # Supprimer les valeurs vides
        all_msisdns = {m for m in all_msisdns if m and str(m).strip()}
        
        if not all_msisdns:
            print("AVERTISSEMENT: Aucun MSISDN trouvé pour l'onglet PB_Format")
            return False
        
        # Créer l'onglet PB_Format
        if 'PB_Format' in wb.sheetnames:
            ws = wb['PB_Format']
        else:
            ws = wb.create_sheet(title="PB_Format")
        
        # Écrire l'en-tête
        headers = ["Donnée brute", "Donnée formatée"]
        ws.append(headers)
        
        # Appliquer le style d'en-tête
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25
        
        # Trier les MSISDN pour un affichage cohérent
        sorted_msisdns = sorted(all_msisdns)
        
        # Formater chaque MSISDN
        for msisdn in sorted_msisdns:
            formatted = format_msisdn(str(msisdn), ref_df)
            ws.append([msisdn, formatted])
        
        print(f"SUCCESS: Onglet PB_Format créé avec {len(sorted_msisdns)} MSISDN")
        return True
        
    except Exception as e:
        print(f"ERREUR: Échec de la création de l'onglet PB_Format - {e}")
        return False


def create_accueil_sheet(wb, tp_df, concat_df):
    """Créer l'onglet Accueil avec les statistiques pour chaque membre du TP
    
    Args:
        wb: Workbook openpyxl
        tp_df: DataFrame du TP
        concat_df: DataFrame concaténé des SMS
    """
    try:
        # Créer l'onglet Accueil (en premier)
        if 'Accueil' in wb.sheetnames:
            ws = wb['Accueil']
        else:
            ws = wb.create_sheet(title="Accueil")
            # Déplacer en premier
            wb.move_endsheet(ws, offset=-len(wb.sheetnames) + 1)
        
        # Écrire l'en-tête
        headers = [
            "Identité",
            "MSISDN",
            "Statut",
            "SMS reçus",
            "Correspondants uniques",
            "Détails correspondants"
        ]
        ws.append(headers)
        
        # Appliquer le style d'en-tête
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
        
        # Analyser les SMS pour chaque membre du TP
        for _, tp_row in tp_df.iterrows():
            msisdn = str(tp_row['MSISDN'])
            identite = tp_row['Identité']
            
            # Filtrer les SMS où ce MSISDN apparaît (APT ou APE)
            sms_filtered = concat_df[
                (concat_df['MSISDN APT'].astype(str) == msisdn) |
                (concat_df['MSISDN APE'].astype(str) == msisdn)
            ]
            
            # Calculer les statistiques
            nb_sms = len(sms_filtered)
            
            # Trouver les correspondants uniques (tous les MSISDN en contact)
            correspondants = set()
            for _, row in sms_filtered.iterrows():
                apt = str(row['MSISDN APT'])
                ape = str(row['MSISDN APE'])
                if apt != msisdn:
                    correspondants.add(apt)
                if ape != msisdn:
                    correspondants.add(ape)
            
            # Filtrer pour ne garder que les correspondants du TP
            tp_msisdns = set(tp_df['MSISDN'].astype(str))
            correspondants_tp = correspondants & tp_msisdns
            correspondants_externes = correspondants - tp_msisdns
            
            nb_correspondants = len(correspondants)
            
            # Déterminer le statut
            if nb_sms > 0:
                statut = "Détecté"
            else:
                statut = "Non détecté"
            
            # Créer la liste des correspondants (pour la colonne Détails)
            details = ", ".join(sorted(correspondants_tp)) if correspondants_tp else "Aucun"
            if correspondants_externes:
                details += f" (+{len(correspondants_externes)} externes)"
            
            # Écrire la ligne
            ws.append([
                identite,
                msisdn,
                statut,
                nb_sms,
                nb_correspondants,
                details
            ])
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 20  # Identité
        ws.column_dimensions['B'].width = 15  # MSISDN
        ws.column_dimensions['C'].width = 12  # Statut
        ws.column_dimensions['D'].width = 12  # SMS reçus
        ws.column_dimensions['E'].width = 20  # Correspondants uniques
        ws.column_dimensions['F'].width = 40  # Détails correspondants
        
        print(f"SUCCESS: Onglet Accueil créé avec statistiques pour {len(tp_df)} membres")
        return True
        
    except Exception as e:
        print(f"ERREUR: Échec de la création de l'onglet Accueil - {e}")
        return False


def route_sms(concat_file, tp_df, output_file):
    """Router les SMS dans les onglets correspondants avec coloration
    
    Returns:
        tuple: (success: bool, concat_df: DataFrame or None)
    """
    try:
        # Charger le fichier concaténé
        df = pd.read_excel(concat_file)
        
        # Formater pour éviter le format scientifique
        df = format_dataframe(df)
        
        # Vérifier les colonnes requises
        required_cols = ['ActionDate', 'MSISDN APT', 'MSISDN APE', 'Preview', 'Traduction']
        for col in required_cols:
            if col not in df.columns:
                print(f"ERREUR: Colonne manquante dans concat - {col}")
                return False, None
        
        # Charger le workbook de sortie
        wb = load_workbook(output_file)
        
        # Créer le mapping MSISDN -> nom d'onglet
        msisdn_to_sheet = {}
        for _, row in tp_df.iterrows():
            sheet_name = f"{row['Identité']} - {row['MSISDN']}"[:31]
            # Trouver le nom réel (peut avoir été modifié pour les doublons)
            for real_name in wb.sheetnames:
                if real_name.startswith(sheet_name):
                    msisdn_to_sheet[row["MSISDN"]] = real_name
                    break
        
        # Créer le set de tous les MSISDN du TP pour vérification
        all_tp_msisdns = set(tp_df['MSISDN'].astype(str))
        
        print(f"Mapping MSISDN -> Onglets: {len(msisdn_to_sheet)} entrées")
        
        # Parser chaque ligne et copier dans les onglets
        for _, row in df.iterrows():
            apt = str(row["MSISDN APT"])
            ape = str(row["MSISDN APE"])
            
            row_data = [
                row["ActionDate"],
                row["MSISDN APT"],
                row["MSISDN APE"],
                row["Preview"],
                row["Traduction"]
            ]
            
            # Écrire dans APT si présent dans TP
            if apt in msisdn_to_sheet:
                sheet_name = msisdn_to_sheet[apt]
                ws = wb[sheet_name]
                new_row = ws.max_row + 1
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=new_row, column=col_idx, value=value)
                
                # Appliquer la coloration pour cette ligne
                apply_coloration(ws, new_row, apt, ape, all_tp_msisdns)
            
            # Écrire dans APE si différent et présent dans TP
            if ape in msisdn_to_sheet and ape != apt:
                sheet_name = msisdn_to_sheet[ape]
                ws = wb[sheet_name]
                new_row = ws.max_row + 1
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=new_row, column=col_idx, value=value)
                
                # Appliquer la coloration pour cette ligne
                apply_coloration(ws, new_row, apt, ape, all_tp_msisdns)
        
        wb.save(output_file)
        print(f"SUCCESS: Routing SMS terminé avec coloration - {output_file}")
        return True, df
        
    except Exception as e:
        print(f"ERREUR: Échec du routing SMS - {e}")
        return False, None


def apply_coloration(ws, row_num, apt, ape, all_tp_msisdns):
    """Appliquer la coloration aux cellules MSISDN de la ligne
    
    Args:
        ws: Worksheet openpyxl
        row_num: Numéro de la ligne
        apt: MSISDN APT de la ligne
        ape: MSISDN APE de la ligne
        all_tp_msisdns: Set de tous les MSISDN du TP
    """
    try:
        # Trouver les colonnes MSISDN APT et MSISDN APE (colonne 2 et 3)
        apt_col = 2  # MSISDN APT
        ape_col = 3  # MSISDN APE
        
        # Extraire le MSISDN de l'onglet à partir du nom
        # Format: "Identité - MSISDN"
        sheet_msisdn = None
        if " - " in ws.title:
            sheet_msisdn = ws.title.split(" - ")[-1].strip()
        
        if sheet_msisdn is None:
            return
        
        # Vérifier et colorer MSISDN APT
        apt_cell = ws.cell(row=row_num, column=apt_col)
        if str(apt) == str(sheet_msisdn):
            # Jaune: correspond au MSISDN de l'onglet
            apt_cell.fill = YELLOW_FILL
        elif str(apt) in all_tp_msisdns:
            # Rouge: autre MSISDN du TP
            apt_cell.fill = RED_FILL
        
        # Vérifier et colorer MSISDN APE
        ape_cell = ws.cell(row=row_num, column=ape_col)
        if str(ape) == str(sheet_msisdn):
            # Jaune: correspond au MSISDN de l'onglet
            ape_cell.fill = YELLOW_FILL
        elif str(ape) in all_tp_msisdns:
            # Rouge: autre MSISDN du TP
            ape_cell.fill = RED_FILL
            
    except Exception as e:
        print(f"AVERTISSEMENT: Échec de la coloration - {e}")


def format_dataframe(df):
    """Formater le DataFrame pour éviter le format scientifique"""
    try:
        # Convertir toutes les colonnes numériques en string si elles contiennent des MSISDN
        for col in df.columns:
            # Vérifier si la colonne contient des numéros de téléphone
            if col in ['MSISDN APT', 'MSISDN APE', 'MSISDN']:
                df[col] = df[col].apply(lambda x: f"{int(x)}" if pd.notna(x) and str(x).replace('.', '').isdigit() else str(x))
            # Pour les autres colonnes numériques, formater sans notation scientifique
            elif pd.api.types.is_numeric_dtype(df[col]):
                df[col] = df[col].apply(lambda x: f"{int(x)}" if pd.notna(x) and x == int(x) else (f"{x:.15f}".rstrip('0').rstrip('.') if pd.notna(x) else x))
        return df
    except Exception as e:
        print(f"AVERTISSEMENT: Échec du formatage - {e}")
        return df


def load_reference_file():
    """Charger le fichier de référence des MSISDN"""
    try:
        if not os.path.exists(REFERENCE_FILE):
            print(f"AVERTISSEMENT: Fichier de référence introuvable - {REFERENCE_FILE}")
            return None
        
        ref_df = pd.read_csv(REFERENCE_FILE)
        print(f"SUCCESS: Fichier de référence chargé ({len(ref_df)} entrées)")
        return ref_df
    except Exception as e:
        print(f"AVERTISSEMENT: Échec du chargement du fichier de référence - {e}")
        return None


def format_msisdn(msisdn, ref_df):
    """Formater un MSISDN selon le fichier de référence
    
    Args:
        msisdn: MSISDN à formater (string)
        ref_df: DataFrame du fichier de référence
    
    Returns:
        str: MSISDN formaté ou "Pas de correspondance"
    """
    try:
        if not isinstance(msisdn, str) or not msisdn.strip():
            return "Pas de correspondance"
        
        # Nettoyer le MSISDN (enlever espaces, +, etc.)
        clean_msisdn = msisdn.strip().replace(" ", "").replace("+", "").replace("-", "").replace("(", "").replace(")", "")
        
        if not clean_msisdn.isdigit():
            return "Pas de correspondance"
        
        # Trouver le préfixe qui correspond
        for _, row in ref_df.iterrows():
            prefix = str(row['E164_prefixe_num'])
            regex_sans_plus = row.get('regex_E164_sans_plus', '')
            
            # Vérifier si le MSISDN commence par ce préfixe
            if clean_msisdn.startswith(prefix):
                # Vérifier avec regex si disponible
                if regex_sans_plus:
                    import re
                    if re.match(regex_sans_plus, clean_msisdn):
                        # Formater en E.164: +<indicatif><numéro>
                        indicatif = str(row['Indicatif'])
                        return f"+{indicatif}{clean_msisdn[len(prefix):]}"
                else:
                    # Formater sans regex
                    indicatif = str(row['Indicatif'])
                    return f"+{indicatif}{clean_msisdn[len(prefix):]}"
        
        # Si aucun préfixe ne correspond, essayer de deviner le format
        # en ajoutant un + devant si ce n'est pas déjà fait
        if clean_msisdn.startswith('0'):
            # Numéro local, on ne peut pas formater sans connaître le pays
            return "Pas de correspondance"
        
        # Essayer de formater comme un numéro international
        if len(clean_msisdn) >= 10:
            # Supposer que c'est déjà un numéro international
            return f"+{clean_msisdn}"
        
        return "Pas de correspondance"
        
    except Exception as e:
        print(f"AVERTISSEMENT: Échec du formatage de {msisdn} - {e}")
        return "Pas de correspondance"


def get_latest_concat():
    """Récupérer le dernier fichier concaténé créé"""
    try:
        if not os.path.exists(CONCAT_DIR):
            return None
        
        files = [f for f in os.listdir(CONCAT_DIR) if f.startswith("SMS_Concat_") and f.endswith(".xlsx")]
        
        if not files:
            return None
        
        # Trier par date de modification
        files.sort(key=lambda x: os.path.getmtime(os.path.join(CONCAT_DIR, x)), reverse=True)
        
        return os.path.join(CONCAT_DIR, files[0])
        
    except Exception as e:
        print(f"ERREUR: Échec de la récupération du dernier concat - {e}")
        return None


def run_pipeline():
    """Exécuter le pipeline complet"""
    global ERROR_LOG_FILE
    
    # Initialiser le logging des erreurs
    ts_start = datetime.now().strftime("%Y%m%d_%H%M%S")
    init_error_logging()
    
    print("=" * 60)
    print("DEBUT DU PIPELINE SMS")
    print("=" * 60)
    log_error("DEBUT DU PIPELINE SMS", "INFO")
    
    # Étape 1: Vérifier/créer les répertoires
    print("\n[1/5] Vérification des répertoires...")
    ensure_directories()
    log_error("Répertoires vérifiés/créés", "INFO")
    
    # Étape 2: Concaténer les fichiers Brutes
    print("\n[2/5] Concaténation des fichiers Brutes...")
    concat_file = build_concat()
    
    if concat_file is None:
        # Si le dossier Brutes est vide, essayer d'utiliser le dernier concat existant
        print("AVERTISSEMENT: Dossier Brutes vide, recherche du dernier concat existant...")
        log_error("Dossier Brutes vide, tentative de bascule sur le dernier concat", "AVERTISSEMENT")
        
        last_concat = get_latest_concat()
        if last_concat is not None:
            print(f"Dernier concat trouvé: {last_concat}")
            log_error(f"Utilisation du dernier concat: {last_concat}", "INFO")
            concat_file = last_concat
            # Ne pas archiver car on utilise un concat existant
            archive_brutes_done = False
        else:
            error_msg = "ERREUR FATALE: Impossible de créer le fichier concaténé et aucun concat existant trouvé"
            print(error_msg)
            log_error(error_msg, "ERREUR_FATALE")
            return False
    else:
        log_error(f"Fichier concaténé créé: {concat_file}", "INFO")
        archive_brutes_done = True
    
    # Étape 3: Archiver les fichiers Brutes (uniquement si on a créé un nouveau concat)
    if archive_brutes_done:
        print("\n[3/5] Archivage des fichiers Brutes...")
        if not archive_brutes():
            print("AVERTISSEMENT: L'archivage a échoué, continuation...")
            log_error("Archivage des fichiers Brutes échoué", "AVERTISSEMENT")
        else:
            log_error("Fichiers Brutes archivés avec succès", "INFO")
    else:
        print("\n[3/5] Archivage des fichiers Brutes: non nécessaire (mode analyse)")
        log_error("Archivage sauté: utilisation d'un concat existant", "INFO")
    
    # Étape 4: Charger le fichier TP
    print("\n[4/5] Chargement du fichier TP...")
    tp_df = load_tp_file()
    
    if tp_df is None:
        error_msg = "ERREUR FATALE: Impossible de charger le fichier TP"
        print(error_msg)
        log_error(error_msg, "ERREUR_FATALE")
        return False
    
    log_error(f"Fichier TP chargé: {len(tp_df)} lignes", "INFO")
    
    # Étape 5: Créer les onglets, routing SMS, statistiques et formatage
    print("\n[5/5] Création des onglets, routing SMS, statistiques et formatage...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(SORTIE_DIR, f"TP_SMS_{ts}.xlsx")
    
    sheet_map = create_tp_sheets(tp_df, output_file)
    
    if sheet_map is None:
        print("ERREUR FATALE: Impossible de créer les onglets TP")
        return False
    
    # Charger le concat pour les statistiques
    concat_df = pd.read_excel(concat_file)
    concat_df = format_dataframe(concat_df)
    
    # Router les SMS
    success, _ = route_sms(concat_file, tp_df, output_file)
    
    if not success:
        print("ERREUR FATALE: Le routing SMS a échoué")
        return False
    
    # Charger le fichier de référence pour le formatage
    ref_df = load_reference_file()
    
    # Créer les onglets Accueil et PB_Format
    wb = load_workbook(output_file)
    
    # Créer l'onglet Accueil avec les statistiques
    if not create_accueil_sheet(wb, tp_df, concat_df):
        print("AVERTISSEMENT: Impossible de créer l'onglet Accueil")
    
    # Créer l'onglet PB_Format
    if ref_df is not None:
        if not create_pb_format_sheet(wb, concat_df, ref_df):
            print("AVERTISSEMENT: Impossible de créer l'onglet PB_Format")
    else:
        print("AVERTISSEMENT: Fichier de référence non disponible, onglet PB_Format non créé")
    
    wb.save(output_file)
    
    print("\n" + "=" * 60)
    print("PIPELINE TERMINE AVEC SUCCES")
    print(f"Fichier de sortie: {output_file}")
    if ERROR_LOG_FILE:
        print(f"Fichier de log: {ERROR_LOG_FILE}")
    print("=" * 60)
    
    log_error("PIPELINE TERMINE AVEC SUCCES", "INFO")
    return True


def run_from_existing_concat():
    """Exécuter à partir du dernier fichier concat existant"""
    global ERROR_LOG_FILE
    
    # Initialiser le logging des erreurs
    ts_start = datetime.now().strftime("%Y%m%d_%H%M%S")
    init_error_logging()
    
    print("=" * 60)
    print("DEBUT DU TRAITEMENT A PARTIR DU DERNIER CONCAT")
    print("=" * 60)
    log_error("DEBUT DU TRAITEMENT A PARTIR DU DERNIER CONCAT", "INFO")
    
    # Récupérer le dernier concat
    print("\n[1/3] Recherche du dernier fichier concaténé...")
    concat_file = get_latest_concat()
    
    if concat_file is None:
        error_msg = f"ERREUR: Aucun fichier concaténé trouvé dans {CONCAT_DIR}"
        print(error_msg)
        print("Exécutez d'abord run_pipeline() pour créer un fichier concaténé")
        log_error(error_msg, "ERREUR")
        return False
    
    print(f"Dernier concat trouvé: {concat_file}")
    log_error(f"Dernier concat utilisé: {concat_file}", "INFO")
    
    # Charger le fichier TP
    print("\n[2/3] Chargement du fichier TP...")
    tp_df = load_tp_file()
    
    if tp_df is None:
        error_msg = "ERREUR FATALE: Impossible de charger le fichier TP"
        print(error_msg)
        log_error(error_msg, "ERREUR_FATALE")
        return False
    
    log_error(f"Fichier TP chargé: {len(tp_df)} lignes", "INFO")
    
    # Créer les onglets et router les SMS
    print("\n[3/3] Création des onglets, routing SMS, statistiques et formatage...")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(SORTIE_DIR, f"TP_SMS_{ts}.xlsx")
    
    sheet_map = create_tp_sheets(tp_df, output_file)
    
    if sheet_map is None:
        print("ERREUR FATALE: Impossible de créer les onglets TP")
        return False
    
    # Charger le concat pour les statistiques
    concat_df = pd.read_excel(concat_file)
    concat_df = format_dataframe(concat_df)
    
    # Router les SMS
    success, _ = route_sms(concat_file, tp_df, output_file)
    
    if not success:
        print("ERREUR FATALE: Le routing SMS a échoué")
        return False
    
    # Charger le fichier de référence pour le formatage
    ref_df = load_reference_file()
    
    # Créer les onglets Accueil et PB_Format
    wb = load_workbook(output_file)
    
    # Créer l'onglet Accueil avec les statistiques
    if not create_accueil_sheet(wb, tp_df, concat_df):
        print("AVERTISSEMENT: Impossible de créer l'onglet Accueil")
    
    # Créer l'onglet PB_Format
    if ref_df is not None:
        if not create_pb_format_sheet(wb, concat_df, ref_df):
            print("AVERTISSEMENT: Impossible de créer l'onglet PB_Format")
    else:
        print("AVERTISSEMENT: Fichier de référence non disponible, onglet PB_Format non créé")
    
    wb.save(output_file)
    
    print("\n" + "=" * 60)
    print("TRAITEMENT TERMINE AVEC SUCCES")
    print(f"Fichier de sortie: {output_file}")
    if ERROR_LOG_FILE:
        print(f"Fichier de log: {ERROR_LOG_FILE}")
    print("=" * 60)
    
    log_error("TRAITEMENT TERMINE AVEC SUCCES", "INFO")
    return True


if __name__ == "__main__":
    # Mode d'emploi
    if len(sys.argv) > 1 and sys.argv[1] == "--from-existing":
        success = run_from_existing_concat()
    else:
        success = run_pipeline()
    
    sys.exit(0 if success else 1)
